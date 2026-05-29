#!/usr/bin/env python3
"""Template-based Excel booking exporter (Phase 1).

Opens the styled source-of-truth template ``templates/booking-template.xlsx``
with openpyxl and writes booking data into the existing cells, preserving the
exact layout: styles, fills, borders, merged cells, column widths, formulas,
RTL view, number formats and sheet names. The workbook structure is NEVER
generated from scratch.

Algorithm (driven entirely by ``templates/mapping.json``):
  1. Load the template (data_only=False so Hijri ``=B``/``=K`` formulas survive).
  2. Build a date -> first-row index from the template's own Gregorian column,
     which fixes the supported year (Phase 1 template = 2026).
  3. For each mapped chalet block (تولوم, سكاي): match the app chalet by name,
     map its first 4 active periods (by ``sort``) to the 4 template slots, and
     CLEAR every slot cell (period/phone/amount) for the whole year — a full
     rebuild that wipes any prior sample data while leaving styles intact.
  4. Re-fill from the included bookings:
       - include status == confirmed only; exclude pending/cancelled/completed
         and any soft-deleted (deleted_at) booking.
       - skip + report: chalet not in template, period beyond first 4 / inactive,
         date not present in the template year.
       - never silently overwrite: bookings are processed earliest-created first,
         the first to claim a slot wins, any later collision is a reported
         conflict (the loser is NOT written).
  5. Save ``exports/bookings-YYYY.xlsx`` and ``exports/bookings-YYYY-report.json``.

Only slot cells (period/phone/amount) are touched. Header, date, chalet-name,
and Hijri-formula cells are never modified.

Usage:
  python scripts/export_bookings_excel.py \
      --input scripts/sample-bookings.json \
      --template templates/booking-template.xlsx \
      --mapping templates/mapping.json \
      --output-dir exports
"""
from __future__ import annotations

import argparse
import datetime as _dt
import json
import os
import sys
from collections import Counter

import openpyxl


def log(msg: str) -> None:
    print(msg, flush=True)


def load_json(path: str):
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def extract_workspace(raw):
    """Accept either {chalets,bookings}, or a full workspace with a .data wrapper."""
    if isinstance(raw, dict) and "data" in raw and isinstance(raw["data"], dict):
        raw = raw["data"]
    chalets = raw.get("chalets") or []
    bookings = raw.get("bookings") or []
    return chalets, bookings


def active_periods_sorted(chalet):
    periods = [p for p in (chalet.get("periods") or []) if bool(p.get("active"))]
    return sorted(periods, key=lambda p: _as_num(p.get("sort"), 0))


def _as_num(v, default=0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def period_text(period, mode):
    label = str(period.get("label") or "").strip()
    start = str(period.get("start") or "").strip()
    end = str(period.get("end") or "").strip()
    if mode == "time_range" and start and end:
        return "من %s الى %s" % (start, end)
    if label:
        return label
    if start and end:
        return "من %s الى %s" % (start, end)
    return ""


def build_date_index(ws, gregorian_col, first_data_row, rows_per_date, max_row):
    """Map ISO date string -> first worksheet row, read from the template itself."""
    index = {}
    years = Counter()
    r = first_data_row
    while r <= max_row:
        cell = ws["%s%d" % (gregorian_col, r)]
        v = cell.value
        d = None
        if isinstance(v, _dt.datetime):
            d = v.date()
        elif isinstance(v, _dt.date):
            d = v
        if d is not None:
            index[d.isoformat()] = r
            years[d.year] += 1
        r += rows_per_date
    return index, years


def clear_slot_cells(ws, block, date_index, slots):
    """Clear period/phone/amount cells for every date row in a block (full rebuild)."""
    for first_row in date_index.values():
        for slot in slots:
            row = first_row + int(slot["row_offset"])
            for key in ("period_col", "phone_col", "amount_col"):
                ws["%s%d" % (slot[key], row)].value = None


def created_key(b):
    """Sort key: earliest created_at wins; missing timestamps fall to the end
    (so a real timestamp always beats a missing one), ties keep input order."""
    c = str(b.get("created_at") or "")
    return c if c else "9999-12-31T23:59:59Z"


def main(argv=None):
    ap = argparse.ArgumentParser(description="Template-based Excel booking exporter (Phase 1)")
    ap.add_argument("--input", required=True, help="Path to bookings JSON (workspace shape)")
    ap.add_argument("--template", default="templates/booking-template.xlsx")
    ap.add_argument("--mapping", default="templates/mapping.json")
    ap.add_argument("--output-dir", default="exports")
    ap.add_argument("--year", type=int, default=None, help="Override output year (default: inferred from template)")
    args = ap.parse_args(argv)

    for path in (args.input, args.template, args.mapping):
        if not os.path.isfile(path):
            log("ERROR: missing required file: %s" % path)
            return 2

    mapping = load_json(args.mapping)
    sheet_name = mapping["sheet"]
    first_data_row = int(mapping["first_data_row"])
    rows_per_date = int(mapping["rows_per_date"])
    max_periods = int(mapping.get("max_periods_per_chalet", 4))
    included_statuses = set(mapping.get("included_statuses", ["confirmed"]))
    phone_missing = mapping.get("phone_missing_value", "محجوز")
    text_mode = mapping.get("period_text", "label_or_time_range")
    text_mode = "time_range" if text_mode == "time_range" else "label"

    raw = load_json(args.input)
    chalets, bookings = extract_workspace(raw)

    # Lookups
    chalet_by_id = {str(c.get("id")): c for c in chalets}
    name_to_chalets = {}
    for c in chalets:
        if c.get("deleted_at"):
            continue
        name_to_chalets.setdefault(str(c.get("name") or "").strip(), []).append(c)

    log("Loading template: %s" % args.template)
    wb = openpyxl.load_workbook(args.template, data_only=False)
    if sheet_name not in wb.sheetnames:
        log("ERROR: sheet %r not found. Sheets: %s" % (sheet_name, wb.sheetnames))
        return 2
    ws = wb[sheet_name]

    # Date index from the template's own Gregorian column (block 0).
    greg0 = mapping["blocks"][0]["gregorian_col"]
    date_index, years = build_date_index(ws, greg0, first_data_row, rows_per_date, ws.max_row)
    if not date_index:
        log("ERROR: no dates found in template column %s" % greg0)
        return 2
    template_year = args.year or (years.most_common(1)[0][0] if years else None)
    log("Template year: %s  (dates indexed: %d)" % (template_year, len(date_index)))

    # Per-block setup: match chalet by name, map periods -> slots, clear slots.
    report_blocks = []
    block_runtime = []  # parallel to mapping['blocks']
    for block in mapping["blocks"]:
        name = block["chalet_name"]
        matched = name_to_chalets.get(name, [])
        chosen = matched[0] if matched else None
        period_to_slot = {}
        overflow_periods = []
        ambiguous_ids = [str(c.get("id")) for c in matched[1:]] if len(matched) > 1 else []
        if chosen is not None:
            actives = active_periods_sorted(chosen)
            for idx, p in enumerate(actives):
                if idx < max_periods and idx < len(block["slots"]):
                    period_to_slot[str(p.get("id"))] = idx
                else:
                    overflow_periods.append({"period_id": str(p.get("id")), "label": p.get("label"), "sort": p.get("sort")})
        # Full rebuild: clear slot cells for this block regardless of matches.
        clear_slot_cells(ws, block, date_index, block["slots"])
        block_runtime.append({
            "block": block,
            "chosen_id": str(chosen.get("id")) if chosen else None,
            "ambiguous_ids": ambiguous_ids,
            "period_to_slot": period_to_slot,
        })
        report_blocks.append({
            "chalet_name": name,
            "matched_chalet_id": str(chosen.get("id")) if chosen else None,
            "ambiguous_duplicate_ids": ambiguous_ids,
            "periods_mapped": period_to_slot,
            "periods_overflow": overflow_periods,
        })

    name_to_runtime = {br["block"]["chalet_name"]: br for br in block_runtime}

    # Process bookings earliest-created first so the earliest wins each slot.
    ordered = sorted(bookings, key=created_key)

    written = []
    excluded = []
    skipped = []
    conflicts = []
    occupied = {}  # (chalet_name, row, period_col) -> booking id that won the slot

    for b in ordered:
        bid = str(b.get("id"))
        status = str(b.get("status") or "")
        if b.get("deleted_at"):
            excluded.append({"id": bid, "reason": "deleted"})
            continue
        if status not in included_statuses:
            excluded.append({"id": bid, "reason": "status_%s" % (status or "unknown")})
            continue

        chalet = chalet_by_id.get(str(b.get("chalet_id")))
        if chalet is None:
            skipped.append({"id": bid, "reason": "chalet_not_found", "chalet_id": b.get("chalet_id")})
            continue
        cname = str(chalet.get("name") or "").strip()
        rt = name_to_runtime.get(cname)
        if rt is None or rt["chosen_id"] is None:
            skipped.append({"id": bid, "reason": "chalet_not_in_template", "chalet": cname})
            continue
        if str(b.get("chalet_id")) in rt["ambiguous_ids"]:
            skipped.append({"id": bid, "reason": "ambiguous_duplicate_chalet_name", "chalet": cname, "chalet_id": b.get("chalet_id")})
            continue

        pid = str(b.get("period_id"))
        slot_idx = rt["period_to_slot"].get(pid)
        if slot_idx is None:
            skipped.append({"id": bid, "reason": "period_not_in_first_4_or_inactive", "chalet": cname, "period_id": pid})
            continue

        date = str(b.get("booking_date") or "")
        first_row = date_index.get(date)
        if first_row is None:
            skipped.append({"id": bid, "reason": "date_not_in_template", "chalet": cname, "date": date})
            continue

        slot = rt["block"]["slots"][slot_idx]
        row = first_row + int(slot["row_offset"])
        key = (cname, row, slot["period_col"])
        if key in occupied:
            conflicts.append({
                "kept_id": occupied[key], "dropped_id": bid,
                "chalet": cname, "date": date, "slot_index": slot_idx,
                "cell": "%s%d" % (slot["period_col"], row),
            })
            continue
        occupied[key] = bid

        chosen = chalet  # the booking's own chalet (== matched chalet)
        period_obj = next((p for p in (chosen.get("periods") or []) if str(p.get("id")) == pid), {})
        ptext = period_text(period_obj, text_mode)
        phone = str(b.get("customer_phone") or "").strip() or phone_missing
        amount = b.get("total")
        try:
            amount = float(amount)
            if amount.is_integer():
                amount = int(amount)
        except (TypeError, ValueError):
            amount = b.get("total")

        ws["%s%d" % (slot["period_col"], row)].value = ptext
        ws["%s%d" % (slot["phone_col"], row)].value = phone
        ws["%s%d" % (slot["amount_col"], row)].value = amount
        written.append({"id": bid, "chalet": cname, "date": date, "slot_index": slot_idx,
                        "cells": {"period": "%s%d" % (slot["period_col"], row),
                                   "phone": "%s%d" % (slot["phone_col"], row),
                                   "amount": "%s%d" % (slot["amount_col"], row)}})

    os.makedirs(args.output_dir, exist_ok=True)
    xlsx_path = os.path.join(args.output_dir, "bookings-%s.xlsx" % template_year)
    report_path = os.path.join(args.output_dir, "bookings-%s-report.json" % template_year)
    wb.save(xlsx_path)

    report = {
        "generated_at": _dt.datetime.now(_dt.timezone.utc).isoformat(),
        "template": args.template,
        "sheet": sheet_name,
        "year": template_year,
        "input": args.input,
        "output_xlsx": xlsx_path,
        "phase": 1,
        "rules": {
            "included_statuses": sorted(included_statuses),
            "max_periods_per_chalet": max_periods,
            "supported_chalets": [b["chalet_name"] for b in mapping["blocks"]],
            "full_rebuild": True,
            "overwrite_conflicts": False,
        },
        "summary": {
            "total_input_bookings": len(bookings),
            "written": len(written),
            "excluded": len(excluded),
            "skipped": len(skipped),
            "conflicts": len(conflicts),
        },
        "blocks": report_blocks,
        "written": written,
        "excluded": excluded,
        "skipped": skipped,
        "conflicts": conflicts,
    }
    with open(report_path, "w", encoding="utf-8") as fh:
        json.dump(report, fh, ensure_ascii=False, indent=2)

    log("Wrote %s" % xlsx_path)
    log("Wrote %s" % report_path)
    log("Summary: written=%d excluded=%d skipped=%d conflicts=%d" % (
        len(written), len(excluded), len(skipped), len(conflicts)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
